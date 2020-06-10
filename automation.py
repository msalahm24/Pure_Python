import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']



    for cell in range(2,sheet.max_row+1):
        our_cell=sheet.cell(cell,3).value
        after_correct=our_cell*0.9
        new_column=sheet.cell(cell,4)
        new_column.value=after_correct

    values=Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)

